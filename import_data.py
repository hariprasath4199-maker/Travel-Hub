#!/usr/bin/env python3
"""
Zalaris Travel Hub — Excel Data Importer
=========================================
Reads 'Travel_Hub_Data_Template.xlsx' and writes data into the input/ JSON files.

Usage:
  python3 import_data.py                          # imports all sheets
  python3 import_data.py --sheet "Users"          # imports only the Users sheet
  python3 import_data.py --sheet "Cost Centres"   # imports only cost centres
  python3 import_data.py --sheet "Locations"       # imports only locations
  python3 import_data.py --dry-run                # preview without writing files

Sheets mapped to files:
  Users           → input/users.txt
  Cost Centres    → input/cost_centres.txt
  Locations       → input/locations.txt
  Travel Requests → input/requests.txt
  Visa Requests   → input/visa_requests.txt
  Ticket Bookings → input/ticket_bookings.txt
"""

import json, os, sys, argparse
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(SCRIPT_DIR, "input")
EXCEL_FILE = os.path.join(SCRIPT_DIR, "Travel_Hub_Data_Template.xlsx")


def read_sheet(wb, sheet_name, skip_note_rows=True):
    """Read a sheet into a list of dicts, skipping note/instruction rows."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found in workbook. Skipping.")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header row is always row 1
    headers = [str(h).replace(" *", "").strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        # Skip note rows (italic/grey instruction rows) — they start with text in col1 and have None in most other cols
        values = list(row)
        non_none = [v for v in values if v is not None and str(v).strip()]
        if len(non_none) <= 1:
            continue  # Skip rows with only 1 or 0 values (note rows or empty)
        # Skip rows where first cell looks like a note (starts with common note patterns)
        first = str(values[0]).strip() if values[0] else ""
        if first.startswith("Format") or first.startswith("Role values") or first.startswith("Department values") or first.startswith("Destination") or first.startswith("Visa Type") or first.startswith("Location ID") or first.startswith("Region"):
            continue

        record = {}
        for i, h in enumerate(headers):
            val = values[i] if i < len(values) else None
            if val is not None:
                if isinstance(val, datetime):
                    record[h] = val.strftime("%Y-%m-%d")
                else:
                    record[h] = str(val).strip() if not isinstance(val, (int, float)) else val
            else:
                record[h] = ""
        data.append(record)
    return data


def import_users(wb, dry_run=False):
    """Users sheet → input/users.txt"""
    raw = read_sheet(wb, "Users")
    if not raw:
        print("  Users: No data found.")
        return
    result = []
    for r in raw:
        first = r.get("First Name", "")
        last = r.get("Last Name", "")
        if not first:
            continue
        name = f"{first} {last}".strip()
        avatar = r.get("Avatar URL", "")
        if not avatar:
            avatar = f"https://ui-avatars.com/api/?name={name.replace(' ', '+')}&background=255dad&color=fff&size=128"
        result.append({
            "id": r.get("User ID", ""),
            "firstName": first,
            "lastName": last,
            "name": name,
            "email": r.get("Email", ""),
            "role": r.get("Role", "APPLICANT"),
            "avatar": avatar,
        })
    write_json("users.txt", result, dry_run)


def import_employees(wb, dry_run=False):
    """Employees sheet → input/employees.txt"""
    raw = read_sheet(wb, "Employees")
    if not raw:
        print("  Employees: No data found.")
        return
    result = []
    for r in raw:
        first = r.get("First Name", "")
        if not first:
            continue
        result.append({
            "employeeId": r.get("Employee ID", ""),
            "firstName": first,
            "lastName": r.get("Last Name", ""),
            "email": r.get("Email", ""),
            "mobile": r.get("Mobile", ""),
            "role": r.get("Role / Title", ""),
            "department": r.get("Department", ""),
            "costCentre": r.get("Cost Centre ID", ""),
            "location": r.get("Location ID", ""),
        })
    write_json("employees.txt", result, dry_run)


def import_cost_centres(wb, dry_run=False):
    """Cost Centres sheet → input/cost_centres.txt"""
    raw = read_sheet(wb, "Cost Centres")
    if not raw:
        print("  Cost Centres: No data found.")
        return
    result = []
    for r in raw:
        if not r.get("Name"):
            continue
        result.append({
            "id": r.get("Cost Centre ID", ""),
            "name": r.get("Name", ""),
            "department": r.get("Department", ""),
            "description": r.get("Description", ""),
        })
    write_json("cost_centres.txt", result, dry_run)


def import_locations(wb, dry_run=False):
    """Locations sheet → input/locations.txt"""
    raw = read_sheet(wb, "Locations")
    if not raw:
        print("  Locations: No data found.")
        return
    result = []
    for r in raw:
        if not r.get("City"):
            continue
        result.append({
            "id": r.get("Location ID", ""),
            "city": r.get("City", ""),
            "country": r.get("Country", ""),
            "office": r.get("Office Name", ""),
            "region": r.get("Region", ""),
        })
    write_json("locations.txt", result, dry_run)


def import_travel_requests(wb, dry_run=False):
    """Travel Requests sheet → input/requests.txt"""
    raw = read_sheet(wb, "Travel Requests")
    if not raw:
        print("  Travel Requests: No data found.")
        return
    result = []
    for i, r in enumerate(raw, 5001):
        first = r.get("First Name", "")
        last = r.get("Last Name", "")
        if not first:
            continue
        name = f"{first} {last}".strip()
        flight = float(r.get("Flight Cost (EUR)", 0) or 0)
        hotel = float(r.get("Hotel Cost (EUR)", 0) or 0)
        other = float(r.get("Other Cost (EUR)", 0) or 0)
        total = flight + hotel + other

        dep = r.get("Departure Date", "")
        ret = r.get("Return Date", "")
        dates_str = ""
        nights = ""
        if dep and ret:
            try:
                d = datetime.strptime(dep, "%Y-%m-%d") if isinstance(dep, str) else dep
                rt = datetime.strptime(ret, "%Y-%m-%d") if isinstance(ret, str) else ret
                if isinstance(d, str): d = datetime.strptime(d, "%Y-%m-%d")
                if isinstance(rt, str): rt = datetime.strptime(rt, "%Y-%m-%d")
                dates_str = f"{d.strftime('%b %d')} - {rt.strftime('%b %d')}"
                n = max(1, (rt - d).days)
                nights = f"{n} night{'s' if n != 1 else ''}"
            except:
                dates_str = f"{dep} - {ret}"
                nights = "—"

        rec = {
            "id": f"REQ-{i}",
            "employeeId": r.get("Employee ID", ""),
            "firstName": first,
            "lastName": last,
            "employeeName": name,
            "role": r.get("Role", ""),
            "destination": r.get("Destination", ""),
            "dates": dates_str,
            "nights": nights,
            "purpose": r.get("Purpose", ""),
            "cost": f"\u20ac{total:,.2f}",
            "department": r.get("Department", ""),
            "mobileNumber": r.get("Mobile Number", "") or None,
            "status": "PENDING",
            "avatar": f"https://ui-avatars.com/api/?name={name.replace(' ', '+')}&background=255dad&color=fff&size=128",
        }
        # Remove None values
        rec = {k: v for k, v in rec.items() if v is not None}
        result.append(rec)
    write_json("requests.txt", result, dry_run)


def import_visa_requests(wb, dry_run=False):
    """Visa Requests sheet → input/visa_requests.txt"""
    raw = read_sheet(wb, "Visa Requests")
    if not raw:
        print("  Visa Requests: No data found.")
        return
    now = datetime.utcnow().isoformat() + "Z"
    result = []
    for i, r in enumerate(raw, 5001):
        first = r.get("First Name", "")
        last = r.get("Last Name", "")
        if not first:
            continue
        name = f"{first} {last}".strip()

        dep = r.get("Travel Date From", "")
        to = r.get("Travel Date To", "")
        days = 1
        if dep and to:
            try:
                d = datetime.strptime(dep, "%Y-%m-%d") if isinstance(dep, str) else dep
                t = datetime.strptime(to, "%Y-%m-%d") if isinstance(to, str) else to
                days = max(1, (t - d).days)
            except:
                pass

        rec = {
            "id": f"VISA-{i}",
            "currentStep": 1,
            "status": "SUBMITTED",
            "employeeId": r.get("Employee ID", ""),
            "firstName": first,
            "lastName": last,
            "employeeName": name,
            "applicantEmail": r.get("Email", ""),
            "applicantMobile": r.get("Mobile Number", "") or None,
            "passportNumber": r.get("Passport Number", ""),
            "employeeRole": r.get("Visa Type", ""),
            "destination": r.get("Destination Country", ""),
            "travelLocation": r.get("Destination Country", ""),
            "travelDateFrom": dep,
            "travelDateTo": to,
            "numberOfDays": days,
            "urgency": r.get("Urgency", "NORMAL") or "NORMAL",
            "purpose": r.get("Purpose", ""),
            "costCentre": r.get("Cost Centre ID", ""),
            "managerName": r.get("Manager Name", ""),
            "managerEmail": r.get("Manager Email", ""),
            "managerComments": r.get("Manager Comments", "") or "",
            "recommendationLetterFile": r.get("Recommendation Letter File", "") or None,
            "employeeAvatar": f"https://ui-avatars.com/api/?name={name.replace(' ', '+')}&background=6750A4&color=fff",
            "workflowHistory": [{
                "step": 1,
                "action": "Visa request submitted by manager",
                "fromStatus": None,
                "toStatus": "SUBMITTED",
                "performedBy": r.get("Manager Name", "Manager"),
                "performedByRole": "MANAGER",
                "timestamp": now,
            }],
            "emailLog": [],
            "createdAt": now,
            "updatedAt": now,
        }
        rec = {k: v for k, v in rec.items() if v is not None}
        result.append(rec)
    write_json("visa_requests.txt", result, dry_run)


def import_ticket_bookings(wb, dry_run=False):
    """Ticket Bookings sheet → input/ticket_bookings.txt"""
    raw = read_sheet(wb, "Ticket Bookings")
    if not raw:
        print("  Ticket Bookings: No data found.")
        return
    now = datetime.utcnow().isoformat() + "Z"
    result = []
    for i, r in enumerate(raw, 5001):
        first = r.get("First Name", "")
        last = r.get("Last Name", "")
        if not first:
            continue
        name = f"{first} {last}".strip()
        rec = {
            "id": f"TB-{i}",
            "visaRequestId": r.get("Visa Request ID", "") or "",
            "employeeId": r.get("Employee ID", ""),
            "firstName": first,
            "lastName": last,
            "applicantName": name,
            "applicantEmail": r.get("Applicant Email", ""),
            "applicantMobile": r.get("Mobile Number", "") or None,
            "managerName": r.get("Manager Name", ""),
            "managerEmail": r.get("Manager Email", ""),
            "destination": r.get("Destination", ""),
            "travelStartDate": r.get("Travel Start Date", ""),
            "travelEndDate": r.get("Travel End Date", ""),
            "purpose": r.get("Purpose", ""),
            "status": "TICKET_REQUESTED",
            "currentStep": 1,
            "itineraryOptions": [],
            "workflowHistory": [{
                "step": 1,
                "action": "Ticket booking request submitted",
                "fromStatus": None,
                "toStatus": "TICKET_REQUESTED",
                "performedBy": r.get("Manager Name", "Unknown"),
                "performedByRole": "MANAGER",
                "timestamp": now,
            }],
            "emailLog": [],
            "createdAt": now,
            "updatedAt": now,
        }
        rec = {k: v for k, v in rec.items() if v is not None}
        result.append(rec)
    write_json("ticket_bookings.txt", result, dry_run)


def write_json(filename, data, dry_run=False):
    filepath = os.path.join(INPUT_DIR, filename)
    if dry_run:
        print(f"  [DRY RUN] {filename}: {len(data)} records would be written")
        for d in data[:2]:
            print(f"    Sample: {json.dumps(d, indent=2)[:200]}...")
    else:
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"  {filename}: {len(data)} records written successfully")


def main():
    parser = argparse.ArgumentParser(description="Import Excel data into Travel Hub")
    parser.add_argument("--sheet", help="Import only a specific sheet", default=None)
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing files")
    parser.add_argument("--file", default=EXCEL_FILE, help="Path to Excel file")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: Excel file not found: {args.file}")
        print(f"Expected at: {EXCEL_FILE}")
        sys.exit(1)

    print(f"\nZalaris Travel Hub — Data Importer")
    print(f"{'=' * 40}")
    print(f"Reading: {args.file}")
    if args.dry_run:
        print(f"Mode: DRY RUN (no files will be modified)\n")
    else:
        print(f"Mode: LIVE IMPORT\n")

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

    sheet_map = {
        "Users": import_users,
        "Employees": import_employees,
        "Cost Centres": import_cost_centres,
        "Locations": import_locations,
        "Travel Requests": import_travel_requests,
        "Visa Requests": import_visa_requests,
        "Ticket Bookings": import_ticket_bookings,
    }

    if args.sheet:
        if args.sheet in sheet_map:
            print(f"Importing: {args.sheet}")
            sheet_map[args.sheet](wb, args.dry_run)
        else:
            print(f"ERROR: Unknown sheet '{args.sheet}'. Available: {', '.join(sheet_map.keys())}")
            sys.exit(1)
    else:
        for name, fn in sheet_map.items():
            print(f"Importing: {name}")
            fn(wb, args.dry_run)

    wb.close()
    print(f"\n{'=' * 40}")
    if args.dry_run:
        print("Dry run complete. No files were modified.")
    else:
        print("Import complete! Restart the backend server to load new data.")
        print("  cd backend && npx tsx src/server.ts")


if __name__ == "__main__":
    main()
